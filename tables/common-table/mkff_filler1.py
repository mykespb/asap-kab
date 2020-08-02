#!/usr/bin/python3

# kab/asap project
# mkff-filler1.py
# Mikhail (myke) Kolodin, 2020
# 2020-08-01 1.3
# fill database mkff from excel file with joint bards archives data

import pymysql
#import pymysql.cursors

import openpyxl as xl
from openpyxl import Workbook

import datetime as dt

# ------------------------------------------

host = 'localhost'
db   = 'mkff'
user = 'bard'
pasw = 'bard'
cset = 'cp1251'

xlfile = 'ff.xlsx'

# ------------------------------------------

print ("Start working...")

con = pymysql.connect(host=host, user=user, password=pasw, db=db, charset=cset)
#con = pymysql.connect(host=host, user=user, password=pasw, db=db, charset=cset, cursorclass=pymysql.cursors.DictCursor)

# ------------------------------------------

print("\n-----------------------------\nSchemes\n")

# sheet1 scheme
ss1 = "number record_code archive_owner digitize device_no file_name no_on_device timer quality song_line1 rem_date_work song_name singer rem_singer song_author music_author rem_music verses_author record_type record_owner archive_code name_coded line1_coded music_coded author_coded verses_coded uuid_record uuid_original etc_auto rem_dt quality_coded dt_exec".split()
ss1nk = {k: v for k, v in enumerate(ss1)}
ss1kn = {v: k for k, v in enumerate(ss1)}
print("ss1nk", ss1nk)
print("ss1kn", ss1kn)

# sheet1 scheme
ss2 = "number fname fext filename fpath ftype fullpath".split()
ss2nk = {k: v for k, v in enumerate(ss2)}
ss2kn = {v: k for k, v in enumerate(ss2)}
print("ss2nk", ss2nk)
print("ss2kn", ss2kn)

# ------------------------------------------

def test1():
    """simply test db connect"""

    global con, cur

    cur.execute("SELECT VERSION() as ver")
    version = cur.fetchone()
#    print("Database version: {}".format(version['ver']))

    cur.execute("select * from main limit 1")
    desc = cur.description
#    print ("---\ndesc:", desc)
    pdesc = [a[0] for a in desc]
    print ("---\npretty desc:", pdesc)

#        pretty desc: ['number', 'record_code', 'archive_owner', 'digitize', 'device_no', 'file_name', 'no_on_device', 'timer', 'quality', 'song_line1', 'rem_date_work', 'song_name', 'singer', 'rem_singer', 'song_author', 'music_author', 'rem_music', 'verses_author', 'record_type', 'record_owner', 'archive_code', 'name_coded', 'line1_coded', 'music_coded', 'author_coded', 'verses_coded', 'uuid_record', 'uuid_original', 'etc_auto', 'rem_dt', 'quality_coded', 'dt_exec']

    cur.execute("select * from arch limit 1")
    desc = cur.description
#    print ("---\ndesc:", desc)
    pdesc = [a[0] for a in desc]
    print ("---\npretty desc:", pdesc)


# ------------------------------------------

def test2_sh1():
    """test xlsx file access"""

    global con, cur

    wb = xl.load_workbook(xlfile, read_only=True)
    print(wb.sheetnames)

    sh1 = wb.worksheets[0]
    de = sh1['A1'].value
    print (f"1st element is {de}")

    # lsit values for 5 lines, OK
    # ~ n = 0
    # ~ for row in sh1.values:
        # ~ n += 1
        # ~ if n>5: break
        # ~ print ("=" * 100, "\nrow", n)
        # ~ for val in row:
            # ~ print (val)

    # list row, OK
    # ~ for row in sh1.iter_rows(min_row=3, max_row=5, max_col=10, values_only=True):
        # ~ print(row)

    # get a value, OK
    row = sh1[3]
    print(row[0].value)

# ------------------------------------------

def sh1_fill():
    """fill db from xlsx sheet1"""

    global con, cur

    print("start conversion...")
    wb = xl.load_workbook(xlfile, read_only=True)
    sh1 = wb.worksheets[0]

    #cur.execute("SELECT VERSION() as ver")
    #version = cur.fetchone()
    #print("Database version: {}".format(version['ver']))

    rown = 0
    for row in sh1.values:
        rown += 1
        #print(f"working with row #{rown}")
        if rown <= 2: continue
        #if rown >= 10: break

#       number = row[0].value
        record_code = row[0]
        archive_owner = row[1]
        digitize = row[2]
        device_no = row[3]
        file_name = ""
#        file_name = row[4]
        no_on_device = row[5]
        timer = row[6]
        quality = row[7]
        song_line1 = row[8]
        rem_date_work = row[9]
        song_name = row[10]
        singer = row[11]
        rem_singer = row[12]
        song_author = row[13]
        music_author = row[14]
        rem_music = row[15]
        verses_author = row[16]
        record_type = row[19]
        record_owner = row[21]
        archive_code = ""
        name_coded = ""
        line1_coded = ""
        music_coded = ""
        author_coded = ""
        verses_coded = ""
        uuid_record = ""
        uuid_original = ""
        etc_auto = ""
        rem_dt = ""
        quality_coded = ""
        dt_exec = row[24]

        #print("test:", record_code, archive_owner, digitize, device_no, file_name, no_on_device, timer, quality, song_line1, rem_date_work, song_name, singer, rem_singer, song_author, music_author, rem_music, verses_author, record_type, record_owner, rem_dt, dt_exec)

        sql = """insert into main (record_code, archive_owner, digitize, device_no, file_name, no_on_device, timer, quality, song_line1, rem_date_work, song_name, singer, rem_singer, song_author, music_author, rem_music, verses_author, record_type, record_owner, rem_dt, dt_exec) values ('%d', '%s', '%s', '%s', '%s', '%s', '%s', '%s', '%s', '%s', '%s', '%s', '%s', '%s', '%s', '%s', '%s', '%s', '%s', '%s', '%s')""" % (record_code, archive_owner, digitize, device_no, file_name, no_on_device, timer, quality, song_line1, rem_date_work, song_name, singer, rem_singer, song_author, music_author, rem_music, verses_author, record_type, record_owner, rem_dt, dt_exec)

        #sql = """insert into `main` (archive_owner) values ('mykespb')"""

        #print("sql:", sql)

        try:
            #print("executing...")
            cur.execute(sql)
            #print ("committing...")
            con.commit()
            #print("OK")
            if rown % 100 == 0:
                print(rown, end=", ")
        except:
            con.rollback()
            print("failed #", rown)


# ------------------------------------------

def sh3_fill():
    """fill db from xlsx sheet1"""

    global con, cur

    print("start conversion...")
    wb = xl.load_workbook(xlfile, read_only=True)
    sh1 = wb.worksheets[2]

    base = sh1["C1"].value
    print("base=", base)

    rown = 0
    for row in sh1.values:
        rown += 1
        # ~ print(f"working with row #{rown}")
        if rown <= 2: continue
        # ~ if rown >= 4: break

        fname = row[0]
        fext = row[4]
        filename = row[1]
        fpath = row[2]
        ftype = row[5]
        fullpath = base + fpath

        sql = """insert into `arch` (fname, fext, filename, fpath, ftype, fullpath) values ('%s', '%s', '%s', '%s', '%s', '%s')""" % (fname, fext, filename, fpath, ftype, fullpath)

        # ~ print("sql:", sql)

        try:
            # ~ print("executing...")
            cur.execute(sql)
            # ~ print ("committing...")
            con.commit()
            # ~ print("OK")
            if rown % 100 == 0:
                print(rown, end=", ")
        except:
            con.rollback()
            print("failed #", rown)

    print ("sheet 3 done")


# ------------------------------------------

def sh0_fill():
    """test fill db"""

    global con, cur

    print("\n-----------------------------\nTest filling...\n")

#    sql = """insert into `main` (archive_owner) values ('mykespb')"""
    sql = "insert into arch (fname) values ('%s')" % ('my-file',)

    print("sql:", sql)

    try:
        print("executing...")
        cur.execute(sql)
        print ("committing...")
        con.commit()
        print("OK")
    except:
        con.rollback()
        print("failed")


# ------------------------------------------

def main():
    """general starter"""

    global con, cur

    print("\n-----------------------------\nProcessing\n")

    with con:
        with con.cursor() as cur:
        #cur = con.cursor()

            #test1()
            #test2_sh1()
            #sh0_fill()
            #sh1_fill()
            sh3_fill()

    print("\n-----------------------------\nStopping\n")


# ------------------------------------------

main()

#con.close()
print ("Work is over.")

# ------------------------------------------
